# -*- coding: utf-8 -*-
"""
laudo_xlsx.py — Geração do laudo de apuração em planilha Excel.

Produz um .xlsx com quatro abas, alinhadas ao foco do relatório:
  1. Resumo        — IR final do espólio, consolidado.
  2. GCAP          — ganho de capital por bem, campos para o GCAP da RFB.
  3. Herdeiros     — base de custo e ganho diferido por herdeiro.
  4. Memoria FR    — memória de cálculo dos fatores de redução.

Convenção de cores (padrão de modelagem financeira):
  azul   = entrada (valor de origem, conferível em documento)
  preto  = fórmula/cálculo
  amarelo de fundo = campo que exige atenção/conferência

Sistema OrgAudi · ORGATEC CONTABILIDADE E AUDITORIA
"""

from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .dominio import Espolio, StatusApuracao
from .motor import MotorApuracao


# --- estilos -------------------------------------------------------------
_AZUL    = Font(name="Calibri", size=10, color="1F7FB8")   # entrada (azul da paleta)
_PRETO   = Font(name="Calibri", size=10, color="1B2733")   # texto body
_BOLD    = Font(name="Calibri", size=10, bold=True, color="12345E")
_TITULO  = Font(name="Calibri", size=12, bold=True, color="12345E")
_CABEC   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FILL_CAB = PatternFill("solid", fgColor="12345E")   # navy
_FILL_EVN = PatternFill("solid", fgColor="EFF4F9")   # zebra par
_FILL_TOT = PatternFill("solid", fgColor="DCE9F4")   # total row
_FILL_ATN = PatternFill("solid", fgColor="FBF1DA")   # pendência/aviso ambar
_FILL_ERR = PatternFill("solid", fgColor="FBF0EF")   # erro/vermelho
_FILL_OK  = PatternFill("solid", fgColor="E3F1E8")   # OK verde
_BORDA = Border(*([Side(style="thin", color="D4DDE6")] * 4))
_MOEDA = '#,##0.00'


def _cab(ws, linha, titulos, larguras):
    for i, (t, w) in enumerate(zip(titulos, larguras), start=1):
        c = ws.cell(row=linha, column=i, value=t)
        c.font = _CABEC
        c.fill = _FILL_CAB
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _BORDA
        ws.column_dimensions[get_column_letter(i)].width = w


def gerar_laudo_xlsx(espolio: Espolio, caminho_saida: str) -> str:
    """Apura o espólio e grava o laudo em .xlsx. Devolve o caminho."""
    motor = MotorApuracao(espolio)
    resumo = motor.apurar()

    wb = Workbook()

    # ===== ABA 1 — RESUMO =================================================
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = f"LAUDO DE APURAÇÃO — {espolio.nome}"
    ws["A1"].font = _TITULO
    ws["A2"] = f"Óbito: {espolio.data_obito}  |  Partilha: {espolio.data_partilha}"
    ws["A2"].font = _PRETO
    ws["A3"] = "Sistema OrgAudi · ORGATEC CONTABILIDADE E AUDITORIA"
    ws["A3"].font = _PRETO

    _cab(ws, 5, ["Bem", "Regime", "Status", "IR do espólio (R$)"],
         [34, 30, 14, 20])
    ln = 6
    primeira_ir = ln
    for b in espolio.bens:
        r = b.resultado
        ws.cell(row=ln, column=1, value=b.identificacao).font = _PRETO
        ws.cell(row=ln, column=2, value=b.regime.value).font = _PRETO
        st = r.get("status")
        ws.cell(row=ln, column=3,
                value=("OK" if st == StatusApuracao.OK else "PENDENTE")
                ).font = _PRETO
        cir = ws.cell(row=ln, column=4)
        ir = r.get("ir_espolio")
        if ir is None:
            cir.value = "pendente"
            cir.fill = _FILL_ATN
        else:
            cir.value = ir
            cir.number_format = _MOEDA
        cir.font = _PRETO
        for col in range(1, 5):
            ws.cell(row=ln, column=col).border = _BORDA
        ln += 1
    ultima_ir = ln - 1
    # total por fórmula (não hardcode)
    ws.cell(row=ln, column=1, value="IR TOTAL DO ESPÓLIO").font = _BOLD
    ctot = ws.cell(row=ln, column=4)
    ctot.value = f"=SUM(D{primeira_ir}:D{ultima_ir})"
    ctot.font = _BOLD
    ctot.number_format = _MOEDA
    ctot.fill = _FILL_TOT
    for col in range(1, 5):
        ws.cell(row=ln, column=col).border = _BORDA
    if not resumo["conclusivo"]:
        ws.cell(row=ln + 2, column=1,
                value=(f"ATENÇÃO: {resumo['bens_pendentes']} bem(ns) "
                       f"pendente(s). O total NÃO é definitivo.")
                ).font = Font(name="Calibri", size=10, bold=True,
                              color="B33A3A")

    # ===== ABA 2 — GCAP ===================================================
    ws2 = wb.create_sheet("GCAP")
    ws2["A1"] = "GANHO DE CAPITAL POR BEM — campos para o GCAP da RFB"
    ws2["A1"].font = _TITULO
    _cab(ws2, 3, ["Bem", "Custo aquisição\n(DIRPF falecido)",
                  "Valor partilha/venda", "Ganho bruto",
                  "Base tributável\n(após fatores)", "IR (R$)",
                  "Fundamento"],
         [28, 18, 18, 16, 18, 14, 40])
    ln = 4
    for b in espolio.bens:
        r = b.resultado
        ws2.cell(row=ln, column=1, value=b.identificacao).font = _PRETO
        # custo é ENTRADA -> azul
        cc = ws2.cell(row=ln, column=2, value=b.custo_aquisicao_dirpf)
        cc.font = _AZUL
        cc.number_format = _MOEDA
        if b.custo_aquisicao_dirpf is None:
            cc.value = "pendente"
            cc.fill = _FILL_ATN
        cv = ws2.cell(row=ln, column=3,
                      value=b.valor_partilha or b.valor_venda)
        cv.font = _AZUL
        cv.number_format = _MOEDA
        # ganho bruto e base — preto (cálculo do motor)
        gb = r.get("ganho_bruto") or r.get("ganho_diferido")
        cgb = ws2.cell(row=ln, column=4, value=gb)
        cgb.font = _PRETO
        cgb.number_format = _MOEDA
        bt = r.get("base_tributavel")
        cbt = ws2.cell(row=ln, column=5, value=bt)
        cbt.font = _PRETO
        cbt.number_format = _MOEDA
        cir = ws2.cell(row=ln, column=6, value=r.get("ir_espolio"))
        cir.font = _PRETO
        cir.number_format = _MOEDA
        fund = (r.get("nota", "") or "")[:200]
        ws2.cell(row=ln, column=7, value=fund).font = _PRETO
        for col in range(1, 8):
            ws2.cell(row=ln, column=col).border = _BORDA
        ln += 1
    ws2.cell(row=ln + 1, column=1,
             value=("Azul = valor de origem (conferir no documento). "
                    "Preto = cálculo. Amarelo = pendência.")
             ).font = Font(name="Calibri", size=9, italic=True, color="51616F")

    # ===== ABA 3 — HERDEIROS =============================================
    ws3 = wb.create_sheet("Herdeiros")
    ws3["A1"] = "IR DOS HERDEIROS — base de custo recebida e ganho diferido"
    ws3["A1"].font = _TITULO
    _cab(ws3, 3, ["Bem", "Base de custo\ndo herdeiro",
                  "Ganho diferido\n(venda futura)", "IR herança hoje",
                  "Observação"],
         [28, 18, 18, 16, 44])
    ln = 4
    for b in espolio.bens:
        r = b.resultado
        if r.get("status") != StatusApuracao.OK:
            continue
        ws3.cell(row=ln, column=1, value=b.identificacao).font = _PRETO
        base = r.get("base_custo_herdeiro")
        cb = ws3.cell(row=ln, column=2, value=base)
        cb.font = _PRETO
        cb.number_format = _MOEDA
        cd = ws3.cell(row=ln, column=3, value=r.get("ganho_diferido"))
        cd.font = _PRETO
        cd.number_format = _MOEDA
        ws3.cell(row=ln, column=4, value="R$ 0,00 (isento)").font = _PRETO
        obs = ("Herança é rendimento isento (Lei 7.713/88 Art. 6º XVI). "
               "O ganho diferido será tributado na venda futura.")
        ws3.cell(row=ln, column=5, value=obs).font = _PRETO
        for col in range(1, 6):
            ws3.cell(row=ln, column=col).border = _BORDA
        ln += 1

    # ===== ABA 4 — MEMORIA FR ============================================
    ws4 = wb.create_sheet("Memoria FR")
    ws4["A1"] = "MEMÓRIA DE CÁLCULO — FATORES DE REDUÇÃO (Lei 11.196/2005)"
    ws4["A1"].font = _TITULO
    ws4["A2"] = ("Conferir contra o GCAP oficial da RFB. Em divergência, "
                 "prevalece o GCAP.")
    ws4["A2"].font = Font(name="Calibri", size=9, italic=True, color="B33A3A")
    ln = 4
    for b in espolio.bens:
        fr = b.resultado.get("fatores_reducao")
        if fr is None:
            continue
        ws4.cell(row=ln, column=1, value=b.identificacao).font = _BOLD
        ln += 1
        if not fr.aplicavel:
            ws4.cell(row=ln, column=1,
                     value=f"  {fr.motivo_inaplicavel}").font = _PRETO
            ln += 2
            continue
        for passo in fr.memoria:
            ws4.cell(row=ln, column=1, value=f"  {passo}").font = _PRETO
            ln += 1
        ln += 1
    ws4.column_dimensions["A"].width = 110

    wb.save(caminho_saida)
    return caminho_saida
